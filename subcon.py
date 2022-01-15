import openpyxl as xl
from utils import *
import subprocess
import os

def del_rows(sheet):
    len_row = sheet.max_row
    index_row = []
    # loop each row in column
    for index in range(6, len_row+1):
        # define emptiness of cell
        val = sheet.cell(index,1)
        if val.value is None:
            # collect indexes of rows
            index_row.append(index)

    # loop each index value
    for row_del in range(len(index_row)):
        sheet.delete_rows(idx=index_row[row_del], amount=1)
        # exclude offset of rows through each iteration
        index_row = list(map(lambda k: k - 1, index_row))


os.chdir("/home/user044/Documents/Eko Putra/New Subcon/")
ooh = "/home/user044/Documents/Eko Putra/OOH/"

while True:
    try:
        file_date = input('Enter OOH Date Release [ex. 1-jan-20]\n ==> ')
        
        source_file = f'{ooh}ooh_{file_date.lower()}_sc.xlsx'
        target_file = 'Subcon Master Data.xlsx'

        wb_src = xl.load_workbook(source_file)
        wb_tgt = xl.load_workbook(target_file)
        sheet = wb_src.active
        dest_sheet = wb_tgt.active

        break
    except FileNotFoundError:
        print('File Not Found, Enter correct filedate-name')


data_input = input_data(sheet)

#Add option if want to add washing column or not
while True:
    try:
        wash_opt = input('Want to add Washing Cost? (y/n) :')
        if wash_opt.lower() == 'y':
            wb_tgt.active = 1
            wb_tgt.remove(wb_tgt.active)
            wb_tgt.remove(wb_tgt['Sheet3'])
            break
        elif wash_opt.lower() == 'n':
            wb_tgt.active = 0
            wb_tgt.remove(wb_tgt.active)
            wb_tgt.remove(wb_tgt['Sheet3'])
            break
    except ValueError:
        print('Enter y/n')

newfile = input('Enter new file name: ')
newfilename = newfile+'.xlsx'

wb_tgt.save(newfilename)
new_wb = xl.load_workbook(newfilename)
new_sheet = new_wb.active

#Putting data from list which already input by user to targetted new user file
n = 0
if len(data_input) == 0:
    print('No data Stored')
else:
    for cell in new_sheet.iter_rows(min_row=4, max_row=3000):
        if cell[3].value == None:
            cell[3].value = data_input[n].value
            n += 1
            if n == len(data_input):
                break

new_wb.save(newfilename)

#Vlookup the rest of the data
len_row = new_sheet.max_row
for data in new_sheet.iter_rows(min_row=4, max_row=len_row):
    cmt = data[3].value
    row_number = data[3].row
    for cell in sheet.iter_rows(min_row=6):
        if cell[1].value == cmt:
            #below coulumn to match column by personal reference
            new_sheet.cell(row=row_number, column=1).value = cell[5].value #Buyer
            new_sheet.cell(row=row_number, column=2).value = cell[3].value #Embro
            new_sheet.cell(row=row_number, column=3).value = cell[4].value #Printing
            new_sheet.cell(row=row_number, column=7).value = cell[10].value #Buyer PO Qty
            date = cell[12].value
            year,month,day = date[-4::],date[3:5],date[:2]
            #option 1
            new_sheet.cell(row=row_number, column=8).value = f"=DATE({year},{month},{day})" #Ex-fty
            #option 2
            #new_sheet.cell(row=row_number, column=8).value = f"{month}/{day}"
            new_sheet.cell(row=row_number, column=10).value = cell[11].value #SR
            new_sheet.cell(row=row_number, column=11).value = cell[6].value #Buyer Style No
            new_sheet.cell(row=row_number, column=12).value = cell[9].value #Washing Type
            new_sheet.cell(row=row_number, column=13).value = cell[15].value #Unit Price
            new_sheet.cell(row=row_number, column=14).value = cell[13].value #CMP Cost
            new_sheet.cell(row=row_number, column=15).value = cell[16].value #CM+NP/PC
            if wash_opt == 'y':
                new_sheet.cell(row=row_number, column=16).value = cell[14].value

new_wb.save(newfilename)
del_rows(new_sheet)

#inserting row
while True:
    add_row = input('Want to insert row between rows? (y/n) :')
    if add_row.lower() == 'y':
        while True:
            try:
                insert_row(new_sheet)
                break
            except ValueError:
                print('Enter Number')
        break

    elif add_row.lower() == 'n':
        break
    else:
        print('Enter (y/n)')

new_wb.save(newfilename)
#add formula required
len_row = new_sheet.max_row
row = 4
for data in new_sheet.iter_rows(min_row=4, max_row=len_row-1):
    buyer = data[0].value
    target_ex_fty = data[column_index_from_string('I')-1]
    if wash_opt == 'y':
        price_06 = data[column_index_from_string('R')-1]
        wash_adj = data[column_index_from_string('Q')-1]
    else:
        price_06 = data[column_index_from_string('P')-1]

    if buyer is not None:
        if wash_opt == 'y':
            wash_adj.value = f'=IF(P{row}>=1,P{row}-0.2,IF(P{row}>=0.76,P{row}-0.15,IF(P{row}>=0.41,P{row}-0.1,P{row}-0.05)))'
        
        target_ex_fty.value = f'=H{row}-4'
        price_06.value = f'=O{row}*0.6'
    row += 1

new_wb.save(newfilename)

len_row = new_sheet.max_row
new_sheet[f'G{len_row}'].value = f'=SUM(G4:G{len_row-1})'

new_sheet.title = 'order list'
new_wb.save(newfilename)
print(f'{newfilename} saved.')

while True:
    openfile = input('Want to open file now ? [y/n]\n=> ')
    if openfile.upper() == 'Y':
        opener = 'libreoffice'
        subprocess.call([opener,newfilename])
        break
    elif openfile.upper() == 'N':
        break
    else:
        print('Enter y or n')