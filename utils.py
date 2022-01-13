from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from typing import Any
from copy import copy

#example result of convert (21/E/SB/C/GIOD/00767 - B to GIOD/767/B)
def convert_cmt(cell_data, number='0'):
    index_second_char = cell_data.value.find(number)
    last_char = cell_data.value[-1::1]
    if last_char.isdigit():
        last_char = 'A'

    first_data_value = cell_data.value[10:15]
    second_data_value = int(cell_data.value[index_second_char:index_second_char+5])
    return first_data_value+str(second_data_value)+"/"+last_char

#Delete defined column
#list tuple follow own reference !!!
def del_columns(worksheet):
    tuple_for_cols = ([3,2],[5,1],[6,2],[9,3],[10,4],[12,7],[13,1],[14,1],[15,2],[16,6],[17,13])
    
    for col in tuple_for_cols:
        worksheet.delete_cols(col[0],col[1])

#unhide all column
def unhide_col(worksheet):
    len_col = worksheet.max_column
    for cols in range(1,len_col+1):
        col = get_column_letter(cols)
        worksheet.column_dimensions[col].hidden = False

#giving index number until max value colomn
def add_number(worksheet):
    len_col = worksheet.max_column
    for rows in range(1):
        for cols in range(1,len_col+1):
            worksheet.cell(2,cols).value = cols

#auto width column
def auto_fit(worksheet):
    len_col = worksheet.max_column
    len_row = worksheet.max_row
    for data_cols in range(1,len_col+1):
        column_width = 0
        letter = get_column_letter(data_cols)
        for data in range(1,len_row+1):
            if worksheet[letter+str(data)].value == None:
                continue
            elif column_width < len(str(worksheet[letter+str(data)].value)):
                column_width = len(str(worksheet[letter+str(data)].value))
            if column_width > 25:
                column_width = 25
        worksheet.column_dimensions[letter].width = column_width+2

#delete empty row value
def del_rows(sheet):
    len_row = sheet.max_row
    index_row = []
    # loop each row in column
    for index in range(6, len_row+1):
        # define emptiness of cell
        val = sheet.cell(index,1)
        if val.value is None or val.value[:1].isdigit() == False:
            # collect indexes of rows
            index_row.append(index)

    # loop each index value
    for row_del in range(len(index_row)):
        sheet.delete_rows(idx=index_row[row_del], amount=1)
        # exclude offset of rows through each iteration
        index_row = list(map(lambda k: k - 1, index_row))


#Set border for Cell
def set_border(style_desc):
    return Border(left=Side(style=style_desc), 
                        right=Side(style=style_desc), 
                        top=Side(style=style_desc), 
                        bottom=Side(style=style_desc))

#Auto border defined range    
def auto_border(sheet,end_row,end_col,style='thin',start_row=1,start_col=1):
    for row in range(start_row,end_row+1):
        for col in range(start_col,end_col+1):
            sheet.cell(row,col).border = set_border(style)

#Take copy data row/column to save in list database
def take_data(sheet,range_copy: Any='Enter Row Number or Collumn Letter'):
    data_copy = []
    if isinstance(range_copy,int):
        source_row = sheet.iter_rows(min_row=range_copy, max_row=range_copy)
        for row in source_row:
            for col in row:
                data_copy.append(col.value)
    else:
        range_copy = column_index_from_string(range_copy)
        source_col = sheet.iter_cols(min_col=range_copy, max_col=range_copy)
        for col in source_col:
            for row in col:
                data_copy.append(row.value)
    return data_copy

#Paste data from database
def put_data(sheet,list_data,paste_as: Any='Type row or col',cell: Any='Cell Destination'):
    cell = coordinate_from_string(cell)
    row_number = cell[1]
    column_letter = column_index_from_string(cell[0])
    if paste_as.lower() == 'row':
        dest_row = sheet.iter_rows(min_row=row_number, max_row=row_number,min_col=column_letter, max_col=len(list_data)+column_letter-1)
        for row in dest_row:
            i = 0
            for cell in row:
                cell.value = list_data[i]
                i += 1
    elif paste_as.lower() == 'col':
        dest_col = sheet.iter_cols(min_col=column_letter, max_col=column_letter,min_row= row_number, max_row=len(list_data)+row_number-1)
        for col in dest_col:
            i = 0
            for cell in col:
                cell.value = list_data[i]
                i += 1

#Function take input data from User -> check database -> save in list
def input_data(sheet):
    print('\nEnter CMT, type finish to end adding')
    data_input = []
    while True:
        input_value = input('Enter CMT No : ')
        if input_value.lower() == 'finish':
            print('Finish input')
            break
        #n = []
        for cell in sheet.iter_rows(min_row=6,max_col=2):
            cell_value = cell[1].value
            if input_value.upper() in cell_value:
                data_input.append(cell[1])
                print(cell_value,' Added')
            
            #below func show messege to check in 1 loop if there is no data, but with side effect long loading
            '''
            else:
                n += 1
                if n == len(list(sheet.iter_rows(min_row=6,max_col=1))):
                    print('No Data, pls enter correct CMT') '''
    return data_input

#insert row as requested user
def insert_row(sheet):
    amount = int(input('How many row in between? : '))
    index_row = []
    for row in range(4,(sheet.max_row)):
        cell = sheet[f'A{row}']
        if cell.value != None:
            index_row.append(row)
    
    for insert in range(len(index_row)):
        sheet.insert_rows(idx=index_row[insert]+1, amount=amount)
        index_row = list(map(lambda k: k + amount, index_row))

    for row in range(4,(sheet.max_row+1)):
        cell = sheet[f'A{row}']
        if cell.value is None:
            for col in range(0,sheet.max_column):
                letter = get_column_letter(col+1)
                sc_cell = sheet[f'{letter}{row-1}']
                new_cell = sheet[f'{letter}{row}']
                #copy format row
                new_cell.font = copy(sc_cell.font)
                new_cell.border = copy(sc_cell.border)
                new_cell.fill = copy(sc_cell.fill)
                new_cell.number_format = copy(sc_cell.number_format)
                new_cell.protection = copy(sc_cell.protection)
                new_cell.alignment = copy(sc_cell.alignment)