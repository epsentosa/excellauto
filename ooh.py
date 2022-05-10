import openpyxl as xl
from utils import *
import os

os.chdir("/home/user044/Documents/Eko Putra/OOH/")

while True:
    try:
        file_date = input('Enter Date Release [ex. 1-jan-20]\n ==> ')

        file1 = f'ooh_{file_date.lower()}.xlsx'
        file2 = f'ooh_{file_date.lower()}_sc.xlsx'

        wb = xl.load_workbook(file1)
        sheet = wb.active
        break
    except FileNotFoundError:
        print('File Not Found, Enter correct filedate-name')

del_columns(sheet)
unhide_col(sheet)
del_rows(sheet)
sheet.insert_cols(2)

len_row = sheet.max_row
#convert CMT
for cell in range(6,len_row+1):
    cell_data = sheet['A'+str(cell)]
    target_value = convert_cmt(cell_data)
    sheet['B'+str(cell)].value = target_value

#Add Border
len_col = sheet.max_column
len_row = sheet.max_row
auto_border(sheet,len_row,len_col,start_row=5)

auto_fit(sheet)

sheet.freeze_panes = "C6"

#Adding Auto Filter
len_col = sheet.max_column
last_col = get_column_letter(len_col)
range_filter = f'A5:{last_col}{len_col}'
sheet.auto_filter.ref = range_filter

print('Operation Success')

wb.save(file2)
os.system(f'libreoffice {file2}')
